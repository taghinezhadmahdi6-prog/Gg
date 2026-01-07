import logging
import asyncio
import io
import base64
import json
import re
from datetime import datetime

from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ApplicationBuilder, ContextTypes, CommandHandler, MessageHandler, filters

import google.generativeai as genai  # ✅ Gemini SDK صحیح
from PIL import Image
from openai import AsyncOpenAI

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# ---------------- تنظیمات ----------------
# 1. توکن ربات تلگرام
TELEGRAM_TOKEN = "8229826436:AAGBM8IxFw6zHqhB38b3OmjqrsDprCfKpPA"

# 2. تنظیمات گوگل جمینی (سرویس اصلی)
GOOGLE_API_KEY = "AIzaSyAuvryviPqsfFi8jdUF7fo9nU-eAAqpP_A"
GEMINI_MODEL_ID = "gemini-flash-latest"  # پیشنهاد: gemini-1.5-flash یا gemini-1.5-pro

# 3. تنظیمات Clarifai (سرویس جایگزین)
CLARIFAI_API_KEY = "c21e5e3be76e452ea4c2ffea19b32d58"
CLARIFAI_BASE_URL = "https://api.clarifai.com/v2/ext/openai/v1"
CLARIFAI_MODEL_ID = "https://clarifai.com/openai/chat-completion/models/o4-mini/versions/efcf58b9be9243ffb6e4032e97a40040"
# ----------------------------------------

# ✅ کانفیگ Gemini
genai.configure(api_key=GOOGLE_API_KEY)

# کلاینت Async برای Clarifai
clarifai_client = AsyncOpenAI(
    api_key=CLARIFAI_API_KEY,
    base_url=CLARIFAI_BASE_URL,
)

# حافظه موقت برای ذخیره عکس‌های کاربران
user_invoices = {}

# حافظه گزارش‌ها: ذخیره آخرین N گزارش برای هر کاربر
user_reports = {}
MAX_REPORTS_PER_USER = 5

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

# ---------- ابزارهای کمکی ----------

PERSIAN_DIGITS_MAP = str.maketrans("0123456789", "۰۱۲۳۴۵۶۷۸۹")

def to_persian_digits(s) -> str:
    return str(s).translate(PERSIAN_DIGITS_MAP)

def encode_image_to_base64(pil_image):
    """تبدیل عکس PIL به رشته Base64 برای ارسال به Clarifai بدون افت کیفیت"""
    buffered = io.BytesIO()
    pil_image.save(buffered, format="JPEG", quality=95)
    return base64.b64encode(buffered.getvalue()).decode('utf-8')

def try_extract_json(text: str):
    """تلاش برای استخراج JSON معتبر از متن خروجی مدل"""
    if not text:
        return None
    text = text.strip()

    # اگر کل خروجی JSON است
    try:
        return json.loads(text)
    except Exception:
        pass

    # تلاش: پیدا کردن اولین بلاک {...}
    m = re.search(r"\{[\s\S]*\}", text)
    if not m:
        return None

    candidate = m.group(0)
    try:
        return json.loads(candidate)
    except Exception:
        return None

def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

def build_txt_report(report_obj: dict, raw_fallback_text: str) -> bytes:
    """
    گزارش TXT با جدول‌بندی ساده.
    اگر report_obj معتبر نبود، متن خام را ذخیره می‌کند.
    """
    if not isinstance(report_obj, dict):
        return (raw_fallback_text or "خروجی خالی بود.").encode("utf-8")

    lines = []
    lines.append("گزارش فاکتورها")
    lines.append("-" * 70)

    currency = report_obj.get("currency", "ریال")
    invoices = report_obj.get("invoices", [])
    grand_total = report_obj.get("grand_total", "")

    header = ["شماره فاکتور", "نام کالا", "تعداد", "قیمت واحد", "قیمت کل"]
    col_widths = [14, 28, 8, 14, 14]

    def fmt_row(cols):
        out = []
        for i, c in enumerate(cols):
            c = "" if c is None else str(c)
            if i == 1 and len(c) > col_widths[i]:
                c = c[:col_widths[i]-1] + "…"
            out.append(c.ljust(col_widths[i]))
        return " | ".join(out)

    lines.append(fmt_row(header))
    lines.append("-" * 70)

    for inv in invoices:
        inv_no = inv.get("invoice_no", "")
        items = inv.get("items", [])
        for it in items:
            name = it.get("name", "")
            qty = it.get("qty", "")
            unit_price = it.get("unit_price", "")
            total_price = it.get("total_price", "")

            lines.append(fmt_row([
                inv_no,
                name,
                to_persian_digits(qty),
                to_persian_digits(unit_price),
                to_persian_digits(total_price),
            ]))

    lines.append("-" * 70)
    lines.append(f"جمع کل نهایی: {to_persian_digits(grand_total)} {currency}")

    notes = report_obj.get("notes")
    if notes:
        lines.append("")
        lines.append("توضیحات:")
        lines.append(str(notes))

    return ("\n".join(lines)).encode("utf-8")

def build_xlsx_report(report_obj: dict, raw_fallback_text: str) -> bytes:
    """
    گزارش اکسل:
    - اگر JSON معتبر بود: شیت Items و Summary
    - اگر نبود: یک شیت RawOutput با متن خام
    """
    wb = Workbook()

    if not isinstance(report_obj, dict):
        ws = wb.active
        ws.title = "RawOutput"
        ws["A1"] = "Model Output (Raw)"
        ws["A1"].font = Font(bold=True)
        ws["A2"] = raw_fallback_text or ""
        ws.column_dimensions["A"].width = 120
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    ws = wb.active
    ws.title = "Items"

    headers = ["Invoice No", "Item Name", "Qty", "Unit Price", "Total Price"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="EEEEEE")
    for i, _h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    invoices = report_obj.get("invoices", [])
    for inv in invoices:
        inv_no = inv.get("invoice_no", "")
        for it in inv.get("items", []):
            ws.append([
                inv_no,
                it.get("name", ""),
                it.get("qty", ""),
                it.get("unit_price", ""),
                it.get("total_price", ""),
            ])

    autosize_columns(ws)

    ws2 = wb.create_sheet("Summary")
    currency = report_obj.get("currency", "ریال")
    ws2["A1"] = "Grand Total"
    ws2["B1"] = report_obj.get("grand_total", "")
    ws2["A2"] = "Currency"
    ws2["B2"] = currency

    ws2["A1"].font = Font(bold=True)
    ws2["A2"].font = Font(bold=True)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 30

    notes = report_obj.get("notes")
    if notes:
        ws2["A4"] = "Notes"
        ws2["A4"].font = Font(bold=True)
        ws2["A5"] = str(notes)
        ws2.column_dimensions["B"].width = 80

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

async def send_report_files(update: Update, txt_bytes: bytes, xlsx_bytes: bytes, txt_name: str, xlsx_name: str):
    await update.message.reply_document(
        document=io.BytesIO(txt_bytes),
        filename=txt_name,
        caption="📄 فایل گزارش متنی (TXT)"
    )
    await update.message.reply_document(
        document=io.BytesIO(xlsx_bytes),
        filename=xlsx_name,
        caption="📊 فایل گزارش اکسل (XLSX)"
    )

def store_user_report(user_id: int, txt_bytes: bytes, xlsx_bytes: bytes, txt_name: str, xlsx_name: str):
    if user_id not in user_reports:
        user_reports[user_id] = []

    user_reports[user_id].insert(0, {
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "txt_bytes": txt_bytes,
        "xlsx_bytes": xlsx_bytes,
        "txt_name": txt_name,
        "xlsx_name": xlsx_name,
    })
    user_reports[user_id] = user_reports[user_id][:MAX_REPORTS_PER_USER]

# ---------- پردازش با سرویس‌ها ----------

async def process_with_gemini(images, prompt):
    """
    ✅ Gemini (google-generativeai)
    اجرای sync در thread جدا برای اینکه ربات هنگ نکند
    """
    contents = [prompt]
    contents.extend(images)  # PIL Image را می‌پذیرد

    def _call():
        model = genai.GenerativeModel(GEMINI_MODEL_ID)
        resp = model.generate_content(contents)
        return resp.text

    return await asyncio.to_thread(_call)

async def process_with_clarifai(images, prompt):
    """پردازش با کلاریفای (OpenAI Compatible)"""
    messages_content = [{"type": "text", "text": prompt}]

    for img in images:
        base64_image = encode_image_to_base64(img)
        messages_content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:image/jpeg;base64,{base64_image}",
                "detail": "high"
            }
        })

    response = await clarifai_client.chat.completions.create(
        model=CLARIFAI_MODEL_ID,
        messages=[
            {"role": "system", "content": "Return only valid JSON. No extra text. Extract invoice items accurately."},
            {"role": "user", "content": messages_content}
        ],
        temperature=0.2,
        max_tokens=2000
    )
    return response.choices[0].message.content

# ---------- هندلرها ----------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_invoices[user_id] = []

    keyboard = [
        [KeyboardButton("✅ محاسبه و گزارش نهایی")],
        [KeyboardButton("📁 ارسال گزارش‌های قبلی")]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(
        "سلام! سیستم حسابداری هوشمند آماده است 🚀\n"
        "📸 عکس‌های فاکتور را بفرستید (تعداد نامحدود).\n"
        "🔚 در آخر دکمه «محاسبه و گزارش نهایی» را بزنید.\n"
        "📁 برای دریافت فایل‌های گزارش قبلی دکمه مربوطه را بزنید.",
        reply_markup=reply_markup
    )

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id not in user_invoices:
        user_invoices[user_id] = []

    msg = await update.message.reply_text("📥 در حال دریافت تصویر با کیفیت اصلی...")

    try:
        photo_file = await update.message.photo[-1].get_file()
        image_bytes = await photo_file.download_as_bytearray()
        img = Image.open(io.BytesIO(image_bytes))

        user_invoices[user_id].append(img)
        count = len(user_invoices[user_id])

        await context.bot.edit_message_text(
            chat_id=update.effective_chat.id,
            message_id=msg.message_id,
            text=f"✅ فاکتور {count} ذخیره شد."
        )
    except Exception as e:
        await update.message.reply_text(f"❌ خطا در دریافت عکس: {e}")

async def send_previous_reports(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    reports = user_reports.get(user_id, [])
    if not reports:
        await update.message.reply_text("❌ هنوز هیچ گزارش قبلی برای شما ذخیره نشده است.")
        return

    await update.message.reply_text(f"📁 در حال ارسال {len(reports)} گزارش قبلی (آخرین‌ها)...")
    for idx, r in enumerate(reports, start=1):
        caption = f"گزارش #{idx} | {r['created_at']}"

        await update.message.reply_document(
            document=io.BytesIO(r["txt_bytes"]),
            filename=r["txt_name"],
            caption=f"📄 {caption}"
        )
        await update.message.reply_document(
            document=io.BytesIO(r["xlsx_bytes"]),
            filename=r["xlsx_name"],
            caption=f"📊 {caption}"
        )

async def process_all_invoices(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if user_id not in user_invoices or not user_invoices[user_id]:
        await update.message.reply_text("❌ عکسی برای پردازش وجود ندارد.")
        return

    images = user_invoices[user_id]
    await update.message.reply_text(f"⏳ در حال پردازش {len(images)} فاکتور...")
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")

    # پرامپت ساختاریافته برای اکسل
    prompt_text = (
        "فقط و فقط یک JSON معتبر برگردان (بدون متن اضافی).\n"
        "این تصاویر فاکتور خرید هستند.\n\n"
        "ساختار دقیق خروجی:\n"
        "{\n"
        '  "currency": "ریال",\n'
        '  "invoices": [\n'
        "    {\n"
        '      "invoice_no": "1",\n'
        '      "items": [\n'
        '        {"name": "نام کالا", "qty": 2, "unit_price": 1000, "total_price": 2000}\n'
        "      ],\n"
        '      "invoice_total": 2000\n'
        "    }\n"
        "  ],\n"
        '  "grand_total": 2000,\n'
        '  "notes": "هر توضیح لازم"\n'
        "}\n\n"
        "اگر بخشی ناخوانا بود مقدار را null بگذار."
    )

    result_text = ""
    source_used = ""

    # 1) Gemini
    try:
        logging.info(f"User {user_id}: Trying Gemini...")
        result_text = await process_with_gemini(images, prompt_text)
        source_used = "Google Gemini ⚡️"
    except Exception as e:
        logging.error(f"Gemini Error: {e}")
        await update.message.reply_text(f"⚠️ جمینی پاسخ نداد (خطا: {e}).\n🔄 سوییچ به سرور کمکی (Clarifai)...")

        # 2) Clarifai
        try:
            logging.info(f"User {user_id}: Trying Clarifai fallback...")
            await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
            result_text = await process_with_clarifai(images, prompt_text)
            source_used = "Clarifai AI 🛡️"
        except Exception as e2:
            logging.error(f"Clarifai Error: {e2}")
            await update.message.reply_text(f"❌ هر دو سرویس با خطا مواجه شدند.\nخطای دوم: {e2}")
            return

    if not result_text:
        await update.message.reply_text("⚠️ خروجی خالی بود.")
        return

    report_obj = try_extract_json(result_text)

    now_tag = datetime.now().strftime("%Y%m%d_%H%M%S")
    txt_name = f"report_{user_id}_{now_tag}.txt"
    xlsx_name = f"report_{user_id}_{now_tag}.xlsx"

    txt_bytes = build_txt_report(report_obj, result_text)
    xlsx_bytes = build_xlsx_report(report_obj, result_text)

    store_user_report(user_id, txt_bytes, xlsx_bytes, txt_name, xlsx_name)

    await update.message.reply_text(f"📊 گزارش نهایی آماده شد (توسط {source_used}).\n📎 فایل‌ها را ارسال می‌کنم...")

    await send_report_files(update, txt_bytes, xlsx_bytes, txt_name, xlsx_name)

    user_invoices[user_id] = []
    await update.message.reply_text("✅ حافظه تصاویر پاک شد. برای فاکتورهای جدید دوباره عکس بفرستید.")

if __name__ == '__main__':
    application = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    application.add_handler(CommandHandler('start', start))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(MessageHandler(filters.Regex(r'^✅'), process_all_invoices))
    application.add_handler(MessageHandler(filters.Regex(r'^📁'), send_previous_reports))

    print("ربات با Gemini اصلاح‌شده + خروجی فایل (XLSX/TXT) + گزارش‌های قبلی روشن شد...")
    application.run_polling()
